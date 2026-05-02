import tkinter as tk
from tkinter import messagebox
from openpyxl import Workbook, load_workbook
import re
import os

file_name = 'data.xlsx'

# Load or create workbook
if os.path.exists(file_name):
    wb = load_workbook(file_name)
    ws = wb.active
else:
    wb = Workbook()
    ws = wb.active
    ws.append(["Name", "Age", "Email", "Phone", "Address"])


def is_duplicate_email(email):
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[2] == email:
            return True
    return False


def save_data():
    name = entry_name.get().strip()
    age = entry_age.get().strip()
    email = entry_email.get().strip()
    phone = entry_phone.get().strip()
    address = entry_address.get("1.0", tk.END).strip()

    # Check empty fields
    if not name or not age or not email or not phone or not address:
        messagebox.showwarning("Warning", "All fields are mandatory.")
        return

    # Validate age
    try:
        age = int(age)
        if age <= 0 or age > 120:
            raise ValueError
    except ValueError:
        messagebox.showwarning("Warning", "Enter a valid age (1–120).")
        return

    # Validate email
    if not re.match(r"[^@]+@[^@]+\.[^@]+", email):
        messagebox.showwarning("Warning", "Invalid email format.")
        return

    # Validate phone (7–15 digits)
    if not re.match(r"^\d{7,15}$", phone):
        messagebox.showwarning("Warning", "Phone must be 7–15 digits.")
        return

    # Check duplicate email
    if is_duplicate_email(email):
        messagebox.showwarning("Warning", "Email already exists.")
        return

    # Save data
    ws.append([name, age, email, phone, address])

    try:
        wb.save(file_name)
    except PermissionError:
        messagebox.showerror("Error", "Close the Excel file before saving.")
        return

    messagebox.showinfo("Success", "Data saved successfully!")

    # Clear fields
    entry_name.delete(0, tk.END)
    entry_age.delete(0, tk.END)
    entry_email.delete(0, tk.END)
    entry_phone.delete(0, tk.END)
    entry_address.delete("1.0", tk.END)


# UI Setup
root = tk.Tk()
root.title("Data Entry Form")
root.geometry("350x300")
root.resizable(False, False)
root.configure(bg="#4B875F")

label_style = {"bg": "#4B875F", "fg": "white"}
entry_style = {"bg": '#D3D3D3', "fg": "black"}

# Name
tk.Label(root, text="Name", **label_style).grid(row=0, column=0, padx=10, pady=5)
entry_name = tk.Entry(root, **entry_style)
entry_name.grid(row=0, column=1, padx=10, pady=5)

# Age
tk.Label(root, text="Age", **label_style).grid(row=1, column=0, padx=10, pady=5)
entry_age = tk.Entry(root, **entry_style)
entry_age.grid(row=1, column=1, padx=10, pady=5)

# Email
tk.Label(root, text="Email", **label_style).grid(row=2, column=0, padx=10, pady=5)
entry_email = tk.Entry(root, **entry_style)
entry_email.grid(row=2, column=1, padx=10, pady=5)

# Phone
tk.Label(root, text="Phone", **label_style).grid(row=3, column=0, padx=10, pady=5)
entry_phone = tk.Entry(root, **entry_style)
entry_phone.grid(row=3, column=1, padx=10, pady=5)

# Address (multiline)
tk.Label(root, text="Address", **label_style).grid(row=4, column=0, padx=10, pady=5)
entry_address = tk.Text(root, height=3, width=22)
entry_address.grid(row=4, column=1, padx=10, pady=5)

# Save Button
tk.Button(root, text="Save", command=save_data,
          bg='#6D8299', fg='white', width=17)\
    .grid(row=5, column=0, columnspan=2, padx=10, pady=15)

root.mainloop()